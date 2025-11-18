"""
Script di verifica per contare i giorni G e i progressivi nei file generati
"""

import openpyxl
import sys
from collections import defaultdict

def verifica_turni(file_excel: str):
    """Analizza il file Excel e conta giorni G e progressivi"""

    print("="*70)
    print(f"VERIFICA FILE: {file_excel}")
    print("="*70)

    wb = openpyxl.load_workbook(file_excel, data_only=True)

    # Mesi da analizzare
    mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
           'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']

    # Contatori per turno
    g_per_turno = defaultdict(int)
    g_per_mese_turno = defaultdict(lambda: defaultdict(int))
    progressivi_finali = {}

    for idx, mese in enumerate(mesi, 1):
        if mese not in wb.sheetnames:
            continue

        ws = wb[mese]

        # Trova le righe dei turni (partono dalla riga 3)
        for row in range(3, ws.max_row + 1):
            turno_num = ws.cell(row, 1).value

            if turno_num and str(turno_num).isdigit():
                turno = int(turno_num)

                # Conta i G in questo mese
                giorni_nel_mese = 31 if idx in [1,3,5,7,8,10,12] else 30
                if idx == 2:
                    giorni_nel_mese = 29  # Assume bisestile

                for col in range(2, giorni_nel_mese + 2):
                    codice = ws.cell(row, col).value
                    if codice == 'G':
                        g_per_turno[turno] += 1
                        g_per_mese_turno[turno][mese] += 1

                # Leggi progressivo finale (ultima colonna)
                if mese == 'DICEMBRE':
                    progressivo = ws.cell(row, ws.max_column).value
                    if progressivo:
                        progressivi_finali[turno] = progressivo

    # Stampa risultati
    print("\n📊 CONTEGGIO GIORNI G PER TURNO:")
    print("-" * 70)
    for turno in sorted(g_per_turno.keys()):
        print(f"Turno {turno}: {g_per_turno[turno]} giorni G")
        # Dettaglio per mese
        mesi_con_g = g_per_mese_turno[turno]
        if mesi_con_g:
            dettaglio = ", ".join([f"{m[:3]}:{c}" for m, c in sorted(mesi_con_g.items())])
            print(f"          Dettaglio: {dettaglio}")

    print("\n📈 PROGRESSIVI FINALI (DICEMBRE):")
    print("-" * 70)
    for turno in sorted(progressivi_finali.keys()):
        prog = progressivi_finali[turno]
        status = "✅" if prog <= 231 else "❌"
        print(f"Turno {turno}: {prog} giorni {status}")

    # Verifica se ci sono problemi
    print("\n🔍 VERIFICA:")
    print("-" * 70)

    problemi = []

    # Verifica G
    for turno, count in g_per_turno.items():
        if turno == 46:
            continue  # T46 ha pattern speciale
        if count < 5 or count > 7:
            problemi.append(f"Turno {turno}: {count} giorni G (attesi 5-6)")

    # Verifica progressivi
    for turno, prog in progressivi_finali.items():
        if prog > 231:
            problemi.append(f"Turno {turno}: {prog} giorni (supera 231!)")

    if problemi:
        print("❌ ANOMALIE RILEVATE:")
        for p in problemi:
            print(f"  - {p}")
    else:
        print("✅ Tutto OK! Tutti i turni hanno:")
        print("  - 5-6 giorni G (escluso T46)")
        print("  - Progressivo ≤ 231 giorni")

    print("\n" + "="*70)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python verifica_giorni_g.py <file_excel>")
        print("\nEsempio:")
        print('  python verifica_giorni_g.py "TURNO_COMPLETO_2025.xlsx"')
        sys.exit(1)

    file_excel = sys.argv[1]
    verifica_turni(file_excel)

"""
Verifica conteggio giorni nel file generato
"""
import openpyxl

file_generato = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\PROGETTI_PYTHON_OFFICE\SHIFT_MANAGERV1.0\TURNO_COMPLETO_2025.xlsx"

wb = openpyxl.load_workbook(file_generato, data_only=True)

mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
        'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']

turni = [41, 42, 43, 44, 45, 47, 48, 49, 50, 51]

print("="*80)
print("VERIFICA CONTEGGIO GIORNI - FILE GENERATO")
print("="*80)

for turno in turni:
    shifts = 0  # A+B+C+G
    ferie = 0   # F*
    riposi = 0  # -
    totale = 0

    for mese in mesi:
        if mese not in wb.sheetnames:
            continue

        ws = wb[mese]

        for row in range(3, ws.max_row + 1):
            if ws.cell(row, 1).value == turno:
                giorni_mese = 31 if mese in ['GENNAIO', 'MARZO', 'MAGGIO', 'LUGLIO', 'AGOSTO', 'OTTOBRE', 'DICEMBRE'] else (30 if mese != 'FEBBRAIO' else 28)

                for col in range(2, giorni_mese + 2):
                    val = ws.cell(row, col).value
                    if val:
                        totale += 1
                        if val in ['A', 'B', 'C', 'G']:
                            shifts += 1
                        elif isinstance(val, str) and val.startswith('F'):
                            ferie += 1
                        elif val == '-':
                            riposi += 1
                break

    print(f"\nTurno {turno}:")
    print(f"  Shifts (A+B+C+G): {shifts}")
    print(f"  Ferie (F*): {ferie}")
    print(f"  Riposi (-): {riposi}")
    print(f"  TOTALE: {totale}")
    print(f"  LAVORATIVI (Shifts+Ferie): {shifts + ferie}")

print("\n" + "="*80)
print("CONFRONTO CON UFFICIALE:")
print("  Ufficiale: 219 shifts + 12 ferie = 231 lavorativi")
print("="*80)

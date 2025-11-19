"""
Script per verificare che l'offset sia corretto tra file generato e ufficiale
"""
import pandas as pd
import openpyxl

# File paths
file_ufficiale = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\TURNO COMPLETO 2025 con modifiche da CCNL.xls"
file_generato = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\PROGETTI_PYTHON_OFFICE\SHIFT_MANAGERV1.0\TURNO_COMPLETO_2025.xlsx"

print("="*80)
print("VERIFICA OFFSET E PATTERN - GENNAIO 2025")
print("="*80)

# Leggi file ufficiale
print("\nLettura file UFFICIALE...")
excel_uff = pd.ExcelFile(file_ufficiale, engine='xlrd')
df_uff = pd.read_excel(excel_uff, sheet_name='GENNAIO', header=1)

# Leggi file generato
print("Lettura file GENERATO...")
wb_gen = openpyxl.load_workbook(file_generato, data_only=True)
ws_gen = wb_gen['GENNAIO']

# Lista turni
turni = [41, 42, 43, 44, 45, 47, 48, 49, 50, 51]

print("\n" + "="*80)
print("CONFRONTO PRIMI 10 GIORNI DI GENNAIO")
print("="*80)

differenze_totali = 0

for turno in turni:
    print(f"\n--- TURNO {turno} ---")

    # Trova riga nel file ufficiale
    riga_uff = df_uff[df_uff.iloc[:, 0] == turno]
    if riga_uff.empty:
        print("  Non trovato nel file ufficiale")
        continue

    # Estrai primi 10 giorni dal file ufficiale
    giorni_uff = []
    for col in range(1, 11):
        val = riga_uff.iloc[0, col]
        if pd.isna(val):
            giorni_uff.append('-')
        else:
            giorni_uff.append(str(val).strip())

    # Trova la riga del turno nel file generato
    riga_gen = None
    for r in range(3, ws_gen.max_row + 1):
        if ws_gen.cell(r, 1).value == turno:
            riga_gen = r
            break

    if riga_gen:
        giorni_gen = []
        for col in range(2, 12):
            val = ws_gen.cell(riga_gen, col).value
            if val is None:
                giorni_gen.append('-')
            else:
                giorni_gen.append(str(val).strip())
    else:
        print("  Non trovato nel file generato")
        continue

    # Confronta
    print(f"  UFFICIALE: {' '.join(giorni_uff)}")
    print(f"  GENERATO:  {' '.join(giorni_gen)}")

    # Evidenzia differenze
    differenze = []
    for i, (u, g) in enumerate(zip(giorni_uff, giorni_gen), 1):
        if u != g:
            differenze.append(f"Giorno {i}: '{u}' vs '{g}'")
            differenze_totali += 1

    if differenze:
        print(f"  DIFFERENZE: {', '.join(differenze)}")
    else:
        print("  OK - Pattern identico!")

print("\n" + "="*80)
if differenze_totali == 0:
    print("SUCCESSO! Tutti i pattern corrispondono perfettamente!")
else:
    print(f"ATTENZIONE: Trovate {differenze_totali} differenze")
print("="*80)

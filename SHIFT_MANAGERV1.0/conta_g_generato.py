"""
Conta i G nel file generato per capire la distribuzione
"""
import openpyxl

file_generato = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\PROGETTI_PYTHON_OFFICE\SHIFT_MANAGERV1.0\TURNO_COMPLETO_2025.xlsx"

wb = openpyxl.load_workbook(file_generato, data_only=True)

mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
        'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']

turni = [41, 42, 43, 44, 45, 47, 48, 49, 50, 51]

print("="*80)
print("DISTRIBUZIONE G - FILE GENERATO")
print("="*80)

g_per_turno_per_mese = {}

for turno in turni:
    g_per_turno_per_mese[turno] = {}

    for mese in mesi:
        if mese not in wb.sheetnames:
            continue

        ws = wb[mese]
        g_count = 0

        for row in range(3, ws.max_row + 1):
            if ws.cell(row, 1).value == turno:
                for col in range(2, 33):
                    if ws.cell(row, col).value == 'G':
                        g_count += 1
                break

        g_per_turno_per_mese[turno][mese] = g_count

# Stampa tabella
header = "Turno | " + " | ".join([m[:3] for m in mesi]) + " | TOT"
print(header)
print("-" * len(header))

for turno in turni:
    row = f"  {turno}  |"
    total = 0
    for mese in mesi:
        count = g_per_turno_per_mese[turno].get(mese, 0)
        total += count
        row += f" {count:3d} |"
    row += f" {total:3d}"
    print(row)

print("\n" + "="*80)
print("CONFRONTO CON FILE UFFICIALE:")
print("="*80)
print("\nUfficiale: 1 G in Gen, Feb, Mar, Apr, Oct per TUTTI")
print("           Extra 1 G in Mag per turni 42,44,45,48,50,51")
print("\nNOTA: I G in Gen nel file generato NON dovrebbero esserci!")
print("="*80)

"""
PREVIEW DIALOG
Finestra di anteprima del calendario turni generato
"""

import customtkinter as ctk
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import calendar

class PreviewDialog(ctk.CTkToplevel):
    """Dialog per visualizzare anteprima calendario"""

    def __init__(self, parent, excel_file: str):
        super().__init__(parent)

        self.excel_file = excel_file
        self.current_month = 1
        self.total_months = 12
        self.workbook = None
        self.data = {}

        # Configurazione finestra
        self.title(f"👁️ Anteprima - {excel_file.split('/')[-1].split('\\')[-1]}")
        self.geometry("1100x700")
        self.minsize(900, 600)

        # Carica dati
        if not self.load_data():
            self.destroy()
            return

        # Crea UI
        self.create_widgets()

        # Mostra primo mese
        self.show_month(1)

        # Centra finestra
        self.center_window()

    def center_window(self):
        """Centra la finestra sullo schermo"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def load_data(self) -> bool:
        """Carica i dati dal file Excel"""
        try:
            self.workbook = load_workbook(self.excel_file, data_only=True)

            # Leggi tutti i fogli
            mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
                   'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']

            for i, mese in enumerate(mesi, 1):
                if mese in self.workbook.sheetnames:
                    ws = self.workbook[mese]
                    self.data[i] = self.parse_sheet(ws)

            return True

        except Exception as e:
            messagebox.showerror(
                "Errore",
                f"Impossibile caricare il file:\n{str(e)}"
            )
            return False

    def parse_sheet(self, ws):
        """Estrae dati da un foglio Excel"""
        data = {
            'header': None,
            'giorni': [],
            'turni': []
        }

        # Header (riga 1)
        if ws.cell(1, 1).value:
            data['header'] = ws.cell(1, 1).value

        # Intestazioni giorni (riga 2)
        max_col = ws.max_column
        for col in range(2, max_col - 2):  # Escludi colonne finali
            val = ws.cell(2, col).value
            if val and str(val).isdigit():
                data['giorni'].append(int(val))
            else:
                break

        # Righe turni (dalla riga 3)
        for row in range(3, ws.max_row + 1):
            turno_num = ws.cell(row, 1).value
            if turno_num and str(turno_num).isdigit():
                turno_data = {
                    'numero': int(turno_num),
                    'giorni': [],
                    'totale': ws.cell(row, max_col - 1).value,
                    'progressivo': ws.cell(row, max_col).value
                }

                # Leggi codici giorni
                for col in range(2, len(data['giorni']) + 2):
                    cell = ws.cell(row, col)
                    turno_data['giorni'].append({
                        'value': cell.value or '-',
                        'bg_color': self.get_cell_color(cell)
                    })

                data['turni'].append(turno_data)

        return data

    def get_cell_color(self, cell) -> str:
        """Estrae il colore di background di una cella"""
        try:
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color.rgb
                if color and isinstance(color, str):
                    # Rimuovi eventuale alpha channel
                    if len(color) == 8:
                        color = color[2:]
                    return f"#{color}"
        except:
            pass
        return None

    def create_widgets(self):
        """Crea tutti i widget"""

        # Header
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(20, 10))

        title_label = ctk.CTkLabel(
            header_frame,
            text="👁️ Anteprima Calendario Turni",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack()

        # Navigation
        nav_frame = ctk.CTkFrame(self, fg_color="transparent")
        nav_frame.pack(fill="x", padx=20, pady=10)

        self.prev_btn = ctk.CTkButton(
            nav_frame,
            text="◀ Mese Precedente",
            command=self.prev_month,
            width=150,
            height=35
        )
        self.prev_btn.pack(side="left", padx=(0, 10))

        self.month_label = ctk.CTkLabel(
            nav_frame,
            text="",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.month_label.pack(side="left", fill="x", expand=True)

        self.next_btn = ctk.CTkButton(
            nav_frame,
            text="Mese Successivo ▶",
            command=self.next_month,
            width=150,
            height=35
        )
        self.next_btn.pack(side="right", padx=(10, 0))

        # Main content (scrollable)
        self.content_frame = ctk.CTkScrollableFrame(self)
        self.content_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Info frame
        info_frame = ctk.CTkFrame(self, fg_color="transparent")
        info_frame.pack(fill="x", padx=20, pady=(0, 10))

        # Legenda colori
        legend_frame = ctk.CTkFrame(info_frame)
        legend_frame.pack(side="left")

        ctk.CTkLabel(
            legend_frame,
            text="Legenda:",
            font=ctk.CTkFont(size=11, weight="bold")
        ).pack(side="left", padx=(10, 5))

        colors = [
            ("🟥 Dom/Festivi", "#FF0000"),
            ("🟦 Sabati", "#4472C4"),
            ("🟨 Ferie", "#FFFFCC"),
            ("🟩 G (no T46)", "#C6EFCE")
        ]

        for text, _ in colors:
            ctk.CTkLabel(
                legend_frame,
                text=text,
                font=ctk.CTkFont(size=10)
            ).pack(side="left", padx=5)

        # Bottone chiudi
        close_btn = ctk.CTkButton(
            info_frame,
            text="Chiudi",
            command=self.destroy,
            width=100,
            height=35
        )
        close_btn.pack(side="right", padx=10, pady=5)

    def show_month(self, month: int):
        """Mostra il mese specificato"""
        self.current_month = month

        # Aggiorna label
        mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
               'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']
        self.month_label.configure(text=f"{mesi[month - 1]} ({month}/12)")

        # Aggiorna bottoni
        self.prev_btn.configure(state="normal" if month > 1 else "disabled")
        self.next_btn.configure(state="normal" if month < 12 else "disabled")

        # Pulisci contenuto precedente
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        # Mostra dati mese
        if month in self.data:
            self.render_month(self.data[month])
        else:
            ctk.CTkLabel(
                self.content_frame,
                text="Dati non disponibili per questo mese",
                font=ctk.CTkFont(size=14)
            ).pack(pady=50)

    def render_month(self, month_data):
        """Renderizza i dati del mese"""

        # Header
        if month_data['header']:
            header_label = ctk.CTkLabel(
                self.content_frame,
                text=month_data['header'],
                font=ctk.CTkFont(size=16, weight="bold")
            )
            header_label.pack(pady=(0, 15))

        # Tabella
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True)

        # Intestazione
        header_row = ctk.CTkFrame(table_frame, fg_color="gray25")
        header_row.pack(fill="x", padx=5, pady=(5, 2))

        # Colonna Turno
        ctk.CTkLabel(
            header_row,
            text="T",
            font=ctk.CTkFont(size=10, weight="bold"),
            width=30
        ).pack(side="left", padx=2)

        # Colonne giorni
        cell_width = max(25, 800 // len(month_data['giorni']))  # Adatta larghezza
        for giorno in month_data['giorni']:
            ctk.CTkLabel(
                header_row,
                text=str(giorno),
                font=ctk.CTkFont(size=9, weight="bold"),
                width=cell_width
            ).pack(side="left", padx=1)

        # Colonne finali
        ctk.CTkLabel(
            header_row,
            text="Tot",
            font=ctk.CTkFont(size=9, weight="bold"),
            width=40
        ).pack(side="left", padx=2)

        ctk.CTkLabel(
            header_row,
            text="Prog",
            font=ctk.CTkFont(size=9, weight="bold"),
            width=40
        ).pack(side="left", padx=2)

        # Righe turni
        for i, turno in enumerate(month_data['turni']):
            # Aggiungi spazio prima di T46
            if turno['numero'] == 46:
                spacer = ctk.CTkFrame(table_frame, height=10, fg_color="transparent")
                spacer.pack(fill="x")

            # Riga turno
            row_color = "gray20" if i % 2 == 0 else "gray15"
            turno_row = ctk.CTkFrame(table_frame, fg_color=row_color)
            turno_row.pack(fill="x", padx=5, pady=1)

            # Numero turno
            ctk.CTkLabel(
                turno_row,
                text=str(turno['numero']),
                font=ctk.CTkFont(size=10, weight="bold"),
                width=30
            ).pack(side="left", padx=2)

            # Celle giorni
            for day_data in turno['giorni']:
                cell_frame = ctk.CTkFrame(turno_row, width=cell_width, height=25)
                cell_frame.pack(side="left", padx=1, pady=1)
                cell_frame.pack_propagate(False)

                # Applica colore background se presente
                if day_data['bg_color']:
                    cell_frame.configure(fg_color=day_data['bg_color'])
                    # Testo nero o bianco in base al background
                    text_color = self.get_contrasting_color(day_data['bg_color'])
                else:
                    text_color = "white"

                cell_label = ctk.CTkLabel(
                    cell_frame,
                    text=str(day_data['value']) if day_data['value'] else '-',
                    font=ctk.CTkFont(size=9),
                    text_color=text_color
                )
                cell_label.pack(expand=True)

            # Totale
            ctk.CTkLabel(
                turno_row,
                text=str(turno['totale']) if turno['totale'] else '',
                font=ctk.CTkFont(size=9),
                width=40
            ).pack(side="left", padx=2)

            # Progressivo
            ctk.CTkLabel(
                turno_row,
                text=str(turno['progressivo']) if turno['progressivo'] else '',
                font=ctk.CTkFont(size=9),
                width=40
            ).pack(side="left", padx=2)

            # Aggiungi spazio dopo T46
            if turno['numero'] == 46:
                spacer = ctk.CTkFrame(table_frame, height=10, fg_color="transparent")
                spacer.pack(fill="x")

    def get_contrasting_color(self, bg_color: str) -> str:
        """Restituisce il colore del testo (nero o bianco) in base al background"""
        try:
            # Rimuovi #
            color = bg_color.replace('#', '')

            # Converti in RGB
            r = int(color[0:2], 16)
            g = int(color[2:4], 16)
            b = int(color[4:6], 16)

            # Calcola luminosità (formula standard)
            luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255

            # Usa nero se il background è chiaro, bianco se è scuro
            return "black" if luminance > 0.5 else "white"

        except:
            return "white"

    def prev_month(self):
        """Va al mese precedente"""
        if self.current_month > 1:
            self.show_month(self.current_month - 1)

    def next_month(self):
        """Va al mese successivo"""
        if self.current_month < 12:
            self.show_month(self.current_month + 1)

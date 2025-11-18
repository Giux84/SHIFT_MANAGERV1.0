"""
GENERATORE TURNI RAFFINERIA
Genera automaticamente i turni dell'anno successivo partendo dal template dell'anno corrente,
mantenendo la continuità del ciclo e applicando la rotazione quinquennale delle ferie.
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import calendar
from typing import Dict, List, Tuple, Optional
import os

# ============================================================================
# COSTANTI E CONFIGURAZIONI
# ============================================================================

# Lista turni
TURNI = [41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51]

# Coppie gemelle (stesso pattern base)
COPPIE = {
    41: 47, 42: 48, 43: 49, 44: 50, 45: 51,
    47: 41, 48: 42, 49: 43, 50: 44, 51: 45
}

# Codici turno
CODICI = {
    'A': 'Mattina',
    'B': 'Pomeriggio',
    'C': 'Notte',
    '-': 'Riposo',
    'G': 'Giornaliero',
    'F': 'Ferie',  # FA, FB, FC, FG
}

# Pattern periodo normale (Gen-Mag, Ott-Dic) - Ciclo 10 giorni
PATTERN_NORMALE = ['B', '-', 'A', 'A', '-', '-', 'C', 'C', '-', 'B']

# Offset per periodo normale
OFFSET_NORMALE = {
    41: 0, 47: 0,
    42: 2, 48: 2,
    43: 4, 49: 4,
    45: 6, 51: 6,
    44: 8, 50: 8,
    46: None  # Turno speciale
}

# Pattern turno 46 periodo normale
PATTERN_46_NORMALE = ['-', 'G', 'G', '-', 'G', 'G', '-', 'G', 'G', '-']

# Pattern periodo estivo (20 Giu - 13 Set circa) - Ciclo 9 giorni
PATTERN_ESTIVO = ['B', '-', 'A', 'A', '-', 'C', 'C', '-', 'B']

# Offset per periodo estivo
OFFSET_ESTIVO = {
    41: 0, 42: 2, 43: 5, 44: 7, 46: 4,
    47: 1, 48: 3, 49: 6, 50: 8
    # 45 e 51 variano in base al periodo ferie
}

# Mesi con G
MESI_CON_G = [1, 2, 3, 4, 10, 11]  # Gen, Feb, Mar, Apr, Ott, Nov

# Festività italiane 2025
FESTIVITA_2025 = [
    (1, 1),    # Capodanno
    (1, 6),    # Epifania
    (4, 20),   # Pasqua
    (4, 21),   # Lunedì dell'Angelo
    (4, 25),   # Liberazione
    (5, 1),    # Festa del Lavoro
    (6, 2),    # Repubblica
    (8, 15),   # Ferragosto
    (11, 1),   # Tutti i Santi
    (12, 8),   # Immacolata
    (12, 25),  # Natale
    (12, 26),  # Santo Stefano
]

# Periodi ferie (date fisse)
PERIODI_FERIE = {
    1: {'inizio': (6, 20), 'fine': (7, 8)},   # 20/06 - 08/07
    2: {'inizio': (7, 8), 'fine': (7, 25)},   # 08/07 - 25/07
    3: {'inizio': (7, 25), 'fine': (8, 11)},  # 25/07 - 11/08
    4: {'inizio': (8, 11), 'fine': (8, 28)},  # 11/08 - 28/08
    5: {'inizio': (8, 28), 'fine': (9, 14)},  # 28/08 - 14/09
    6: {'inizio': (9, 14), 'fine': (9, 30)}   # 14/09 - 30/09 (solo T-46)
}

# Assegnazione ferie 2024
FERIE_2024 = {
    1: [45, 51],  # Periodo 1
    2: [42, 48],  # Periodo 2
    3: [41, 47],  # Periodo 3
    4: [44, 50],  # Periodo 4
    5: [43, 49],  # Periodo 5
    6: [46]       # Periodo 6
}

# Assegnazione ferie 2025 (rotazione 1→3→5→2→4)
# Formula: vecchio_periodo → nuovo_periodo
# 1→3, 2→4, 3→5, 4→1, 5→2, 6→6
FERIE_2025 = {
    1: [44, 50],  # Era periodo 4 nel 2024
    2: [43, 49],  # Era periodo 5 nel 2024
    3: [45, 51],  # Era periodo 1 nel 2024
    4: [42, 48],  # Era periodo 2 nel 2024
    5: [41, 47],  # Era periodo 3 nel 2024
    6: [46]       # Sempre periodo 6
}

# ============================================================================
# FUNZIONI UTILITY
# ============================================================================

def is_weekend(data: date) -> bool:
    """Verifica se la data è sabato o domenica"""
    return data.weekday() in [5, 6]

def is_festivo(data: date) -> bool:
    """Verifica se la data è festività"""
    return (data.month, data.day) in FESTIVITA_2025

def is_weekend_or_festivo(data: date) -> bool:
    """Verifica se la data è weekend o festività"""
    return is_weekend(data) or is_festivo(data)

def get_date_range(anno: int) -> List[date]:
    """Genera lista di tutte le date dell'anno"""
    primo_giorno = date(anno, 1, 1)
    ultimo_giorno = date(anno, 12, 31)
    giorni = []

    giorno_corrente = primo_giorno
    while giorno_corrente <= ultimo_giorno:
        giorni.append(giorno_corrente)
        giorno_corrente += timedelta(days=1)

    return giorni

def is_in_periodo_ferie(data: date, periodo: int, anno: int) -> bool:
    """Verifica se la data è nel periodo ferie specificato"""
    if periodo not in PERIODI_FERIE:
        return False

    inizio_mese, inizio_giorno = PERIODI_FERIE[periodo]['inizio']
    fine_mese, fine_giorno = PERIODI_FERIE[periodo]['fine']

    data_inizio = date(anno, inizio_mese, inizio_giorno)
    data_fine = date(anno, fine_mese, fine_giorno)

    return data_inizio <= data < data_fine

# ============================================================================
# CLASSE PRINCIPALE
# ============================================================================

class GeneratoreTurni:
    """Genera i turni per l'anno successivo"""

    def __init__(self, file_template: str, anno_target: int):
        self.file_template = file_template
        self.anno_template = anno_target - 1
        self.anno_target = anno_target
        self.calendario = {}  # {data: {turno: codice}}
        self.template_data = None
        self.offset_iniziale = {}
        # Contatore G per mese per ogni turno {turno: {mese: conteggio}}
        self.g_counter = {turno: {} for turno in TURNI}

    def leggi_template(self) -> bool:
        """Legge il file template dell'anno precedente"""
        try:
            print(f"Lettura template {self.anno_template}...")

            # Prova con diversi engine
            try:
                self.template_data = pd.ExcelFile(self.file_template, engine='xlrd')
            except:
                try:
                    self.template_data = pd.ExcelFile(self.file_template, engine='openpyxl')
                except Exception as e:
                    print(f"Errore lettura file: {e}")
                    return False

            # Estrai i pattern dall'ultimo mese per calcolare la continuità
            ultimo_mese = 12
            nome_foglio = self._get_nome_foglio(ultimo_mese)

            if nome_foglio not in self.template_data.sheet_names:
                print(f"Foglio {nome_foglio} non trovato")
                return False

            df = pd.read_excel(self.template_data, sheet_name=nome_foglio, header=1)

            # Calcola offset per il 31 dicembre
            self._calcola_offset_iniziale(df, ultimo_mese)

            print("Template letto con successo")
            return True

        except Exception as e:
            print(f"Errore nella lettura del template: {e}")
            return False

    def _get_nome_foglio(self, mese: int) -> str:
        """Restituisce il nome del foglio per il mese"""
        nomi_mesi = ['', 'GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO',
                     'GIUGNO', 'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE',
                     'NOVEMBRE', 'DICEMBRE']
        return nomi_mesi[mese]

    def _calcola_offset_iniziale(self, df: pd.DataFrame, mese: int):
        """Calcola l'offset iniziale per ogni turno basandosi sull'ultimo giorno del template"""
        ultimo_giorno = calendar.monthrange(self.anno_template, mese)[1]

        # Il 2024 è bisestile (366 giorni)
        giorni_anno = 366 if calendar.isleap(self.anno_template) else 365

        for turno in TURNI:
            if turno == 46:
                self.offset_iniziale[46] = 0  # Il turno 46 è speciale
                continue

            # Trova la riga del turno
            turno_row = df[df.iloc[:, 0] == turno]
            if turno_row.empty:
                print(f"Attenzione: turno {turno} non trovato nel template")
                self.offset_iniziale[turno] = OFFSET_NORMALE.get(turno, 0)
                continue

            # L'offset normale per periodo normale
            offset_base = OFFSET_NORMALE.get(turno, 0)

            # Calcola la nuova posizione dopo 366 giorni
            # Per ciclo 10 giorni
            self.offset_iniziale[turno] = (offset_base + giorni_anno) % 10

        print(f"Offset iniziali calcolati: {self.offset_iniziale}")

    def genera_calendario(self):
        """Genera il calendario completo per l'anno target"""
        print(f"\nGenerazione calendario {self.anno_target}...")

        date_anno = get_date_range(self.anno_target)

        for data in date_anno:
            self.calendario[data] = {}

            # Determina il periodo
            periodo = self._determina_periodo(data)

            for turno in TURNI:
                codice = self._genera_codice_giorno(turno, data, periodo)
                self.calendario[data][turno] = codice

        print("Calendario base generato")

    def _determina_periodo(self, data: date) -> str:
        """Determina in quale periodo si trova la data"""
        mese = data.month
        giorno = data.day

        # Periodo estivo: circa 20 giugno - 13 settembre
        if (mese == 6 and giorno >= 20) or (mese == 7) or (mese == 8) or (mese == 9 and giorno <= 13):
            return 'estivo'

        # Transizione entrata: 16-19 giugno (prima dell'inizio ferie)
        if mese == 6 and 16 <= giorno < 20:
            return 'transizione_entrata'

        # Transizione uscita: 14-22 settembre
        if mese == 9 and 14 <= giorno <= 22:
            return 'transizione_uscita'

        # Periodo normale: resto dell'anno
        return 'normale'

    def _genera_codice_giorno(self, turno: int, data: date, periodo: str) -> str:
        """Genera il codice per un turno in una specifica data"""

        # Calcola il giorno dell'anno (0-indexed)
        primo_gennaio = date(data.year, 1, 1)
        giorni_trascorsi = (data - primo_gennaio).days

        if periodo == 'normale':
            return self._codice_periodo_normale(turno, giorni_trascorsi, data)
        elif periodo == 'estivo':
            return self._codice_periodo_estivo(turno, giorni_trascorsi, data)
        else:
            # Transizioni: usa pattern normale come base
            return self._codice_periodo_normale(turno, giorni_trascorsi, data)

    def _codice_periodo_normale(self, turno: int, giorni_trascorsi: int, data: date) -> str:
        """Genera codice per periodo normale (ciclo 10 giorni)"""

        if turno == 46:
            # Turno 46 speciale
            posizione = (self.offset_iniziale[46] + giorni_trascorsi) % 10
            return PATTERN_46_NORMALE[posizione]

        offset = self.offset_iniziale.get(turno, OFFSET_NORMALE.get(turno, 0))
        posizione = (offset + giorni_trascorsi) % 10
        codice_base = PATTERN_NORMALE[posizione]

        # Gestione G: sostituisce il '-' in posizione 4 (DOPO le due mattine AA)
        # Pattern: B - A A - - C C - B
        #          0 1 2 3 4 5 6 7 8 9
        #                  ^ qui (posizione 4, primo riposo dopo AA)
        # LIMITE: massimo 1 G al mese, solo nei mesi con G e MAI in weekend/festivi
        if posizione == 4 and data.month in MESI_CON_G:
            if codice_base == '-' and not is_weekend_or_festivo(data):
                # Verifica se abbiamo già messo 1 G in questo mese per questo turno
                mese = data.month
                if self.g_counter[turno].get(mese, 0) < 1:
                    # Aggiungi G e incrementa contatore
                    self.g_counter[turno][mese] = self.g_counter[turno].get(mese, 0) + 1
                    return 'G'

        return codice_base

    def _codice_periodo_estivo(self, turno: int, giorni_trascorsi: int, data: date) -> str:
        """Genera codice per periodo estivo (ciclo 9 giorni)"""

        # Calcola giorni dall'inizio dell'estate (20 giugno)
        inizio_estate = date(data.year, 6, 20)
        giorni_da_inizio_estate = (data - inizio_estate).days

        if giorni_da_inizio_estate < 0:
            # Siamo prima dell'inizio estate, usa pattern normale
            return self._codice_periodo_normale(turno, giorni_trascorsi, data)

        if turno in OFFSET_ESTIVO:
            offset = OFFSET_ESTIVO[turno]
            posizione = (offset + giorni_da_inizio_estate) % 9
            return PATTERN_ESTIVO[posizione]
        else:
            # Per turni 45 e 51 (variano in base alle ferie)
            # Usa un offset di default
            posizione = giorni_da_inizio_estate % 9
            return PATTERN_ESTIVO[posizione]

    def applica_ferie(self):
        """Applica le ferie secondo FERIE_2025"""
        print("\nApplicazione ferie...")

        for periodo, turni_in_ferie in FERIE_2025.items():
            print(f"  Periodo {periodo}: turni {turni_in_ferie}")

            for turno in turni_in_ferie:
                self._applica_ferie_turno(turno, periodo)

        print("Ferie applicate")

    def bilancia_giorni_lavorativi(self, target: int = 231):
        """Bilancia i giorni lavorativi per ogni turno per raggiungere esattamente il target"""
        print(f"\nBilanciamento giorni lavorativi (target: {target})...")

        for turno in TURNI:
            if turno == 46:
                continue  # Il turno 46 ha regole speciali, non bilanciare

            # Conta giorni lavorativi attuali
            giorni_attuali = 0
            for data, codici in self.calendario.items():
                codice = codici[turno]
                if codice in ['A', 'B', 'C', 'G'] or codice.startswith('F'):
                    giorni_attuali += 1

            differenza = target - giorni_attuali

            if differenza > 0:
                # Servono più giorni: aggiungi G extra
                print(f"  Turno {turno}: {giorni_attuali} giorni, aggiungo {differenza} G")
                self._aggiungi_g_extra(turno, differenza)
            elif differenza < 0:
                # Troppi giorni: rimuovi alcuni G
                print(f"  Turno {turno}: {giorni_attuali} giorni, rimuovo {-differenza} G")
                self._rimuovi_g_extra(turno, -differenza)
            else:
                print(f"  Turno {turno}: {giorni_attuali} giorni OK")

        print("Bilanciamento completato")

    def _aggiungi_g_extra(self, turno: int, quanti: int):
        """Aggiunge G extra nei mesi disponibili per raggiungere il target"""
        aggiunti = 0

        # Prima, conta quanti G ci sono già per mese
        g_per_mese = {}
        for data, codici in self.calendario.items():
            if codici[turno] == 'G':
                mese = data.month
                g_per_mese[mese] = g_per_mese.get(mese, 0) + 1

        # Priorità: MAGGIO (mese 5) per G extra, poi gli altri mesi
        # Mesi disponibili: 1, 2, 3, 4, 5, 10, 11 (ma 5=MAGGIO ha priorità)
        mesi_priorita = [5, 1, 2, 3, 4, 10, 11]  # Maggio prima!

        for mese_target in mesi_priorita:
            if aggiunti >= quanti:
                break

            # Salta se questo mese ha già un G (limite 1 per mese, eccetto maggio)
            if mese_target != 5 and g_per_mese.get(mese_target, 0) >= 1:
                continue

            # Cerca riposi disponibili in questo mese
            for data in sorted(self.calendario.keys()):
                if aggiunti >= quanti:
                    break

                if data.month != mese_target:
                    continue

                codice = self.calendario[data][turno]

                # Converti riposi in G
                if codice == '-' and not is_weekend_or_festivo(data):
                    # Verifica il pattern: deve essere dopo AA
                    primo_gennaio = date(data.year, 1, 1)
                    giorni_trascorsi = (data - primo_gennaio).days
                    offset = self.offset_iniziale.get(turno, OFFSET_NORMALE.get(turno, 0))
                    posizione = (offset + giorni_trascorsi) % 10

                    # Aggiungi G solo in posizione 4 o 5 (riposi dopo AA)
                    if posizione in [4, 5]:
                        self.calendario[data][turno] = 'G'
                        aggiunti += 1
                        # Aggiorna contatore
                        g_per_mese[mese_target] = g_per_mese.get(mese_target, 0) + 1
                        # Se non siamo a maggio, esci dopo 1 G per mese
                        if mese_target != 5:
                            break

    def _rimuovi_g_extra(self, turno: int, quanti: int):
        """Rimuove G extra per raggiungere il target"""
        rimossi = 0

        # Cerca G da rimuovere (in ordine inverso per togliere gli ultimi)
        for data in sorted(self.calendario.keys(), reverse=True):
            if rimossi >= quanti:
                break

            if data.month not in MESI_CON_G:
                continue

            codice = self.calendario[data][turno]

            # Converti G in riposi
            if codice == 'G':
                self.calendario[data][turno] = '-'
                rimossi += 1

    def _applica_ferie_turno(self, turno: int, periodo: int):
        """Applica le ferie a un turno per un periodo specifico"""

        # Se il turno ha un gemello, il gemello inizia 1 giorno prima
        turno_gemello = COPPIE.get(turno)

        giorni_ferie = 0
        data_inizio = None

        for data, codici in sorted(self.calendario.items()):
            if is_in_periodo_ferie(data, periodo, self.anno_target):

                if data_inizio is None:
                    data_inizio = data

                codice_corrente = codici[turno]

                # Conta solo giorni lavorativi (non riposi)
                if codice_corrente in ['A', 'B', 'C', 'G']:

                    # Determina il codice ferie basato sul turno corrente
                    if turno == 46:
                        codice_ferie = 'FG'
                    else:
                        codice_ferie = f"F{codice_corrente}"

                    self.calendario[data][turno] = codice_ferie
                    giorni_ferie += 1

                    # Gestione gemello (1 giorno prima)
                    if turno_gemello and giorni_ferie == 1:
                        # Il primo giorno di ferie del gemello è 1 giorno prima
                        data_gemello = data - timedelta(days=1)
                        if data_gemello in self.calendario:
                            codice_gemello = self.calendario[data_gemello][turno_gemello]
                            if codice_gemello in ['A', 'B', 'C', 'G']:
                                if turno_gemello == 46:
                                    self.calendario[data_gemello][turno_gemello] = 'FG'
                                else:
                                    self.calendario[data_gemello][turno_gemello] = f"F{codice_gemello}"

                    if giorni_ferie >= 12:
                        break

    def verifica_copertura(self) -> Dict[str, List]:
        """Verifica che ogni giorno abbia copertura 2-2-2 (2 turni A, 2 B, 2 C)"""
        print("\nVerifica copertura 2-2-2...")

        anomalie = []
        statistiche = {'ok': 0, 'anomalie': 0}

        for data in sorted(self.calendario.keys()):
            conteggi = {'A': 0, 'B': 0, 'C': 0, 'G': 0}

            for turno, codice in self.calendario[data].items():
                # Ignora riposi e ferie
                if codice in ['A', 'B', 'C']:
                    conteggi[codice] += 1
                elif codice == 'G':
                    conteggi['G'] += 1

            # Verifica 2-2-2 per A, B, C
            if conteggi['A'] != 2 or conteggi['B'] != 2 or conteggi['C'] != 2:
                anomalie.append({
                    'data': data,
                    'A': conteggi['A'],
                    'B': conteggi['B'],
                    'C': conteggi['C'],
                    'G': conteggi['G']
                })
                statistiche['anomalie'] += 1
            else:
                statistiche['ok'] += 1

        print(f"  Giorni OK: {statistiche['ok']}")
        print(f"  Anomalie: {statistiche['anomalie']}")

        if anomalie:
            print(f"\n  Prime 10 anomalie:")
            for anomalia in anomalie[:10]:
                print(f"    {anomalia['data']}: A={anomalia['A']}, B={anomalia['B']}, C={anomalia['C']}, G={anomalia['G']}")

        return {
            'anomalie': anomalie,
            'statistiche': statistiche
        }

    def genera_excel(self, output_path: str):
        """Genera il file Excel output"""
        print(f"\nGenerazione file Excel: {output_path}")

        wb = Workbook()
        wb.remove(wb.active)  # Rimuovi il foglio di default

        # Crea un foglio per ogni mese
        for mese in range(1, 13):
            self._crea_foglio_mese(wb, mese)

        # Salva il file
        wb.save(output_path)
        print(f"File salvato: {output_path}")

    def _crea_foglio_mese(self, wb: Workbook, mese: int):
        """Crea un foglio per il mese specificato"""

        nomi_mesi = ['', 'GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO',
                     'GIUGNO', 'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE',
                     'NOVEMBRE', 'DICEMBRE']

        nome_foglio = nomi_mesi[mese]
        ws = wb.create_sheet(nome_foglio)

        # Giorni del mese
        giorni_mese = calendar.monthrange(self.anno_target, mese)[1]

        # Riga 1: Titolo
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=giorni_mese + 4)
        ws['A1'] = f"{nome_foglio} {self.anno_target}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Riga 2: Intestazioni
        ws['A2'] = 'Turno'
        for giorno in range(1, giorni_mese + 1):
            ws.cell(row=2, column=giorno + 1, value=giorno)

        ws.cell(row=2, column=giorni_mese + 2, value='Turno')
        ws.cell(row=2, column=giorni_mese + 3, value='gg.')
        ws.cell(row=2, column=giorni_mese + 4, value='Progr.')

        # Formattazione intestazioni
        for col in range(1, giorni_mese + 5):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Righe dati turni
        riga = 3

        for turno in TURNI:
            # Aggiungi 2 righe bianche PRIMA del turno 46
            if turno == 46:
                riga += 2

            ws.cell(row=riga, column=1, value=turno)

            giorni_lavorativi_mese = 0

            # Popola i giorni del mese
            for giorno in range(1, giorni_mese + 1):
                data = date(self.anno_target, mese, giorno)
                codice = self.calendario.get(data, {}).get(turno, '-')

                cell = ws.cell(row=riga, column=giorno + 1, value=codice)
                cell.alignment = Alignment(horizontal='center')

                # Conteggio giorni lavorativi
                if codice in ['A', 'B', 'C', 'G'] or codice.startswith('F'):
                    giorni_lavorativi_mese += 1

                # Formattazione celle (passa anche il turno per gestire G del turno 46)
                self._formatta_cella(cell, codice, data, turno)

            # Colonne finali
            cell_turno = ws.cell(row=riga, column=giorni_mese + 2, value=turno)
            cell_gg = ws.cell(row=riga, column=giorni_mese + 3, value=giorni_lavorativi_mese)

            # Calcola progressivo PER TURNO (non globale)
            progressivo_turno = self._calcola_progressivo_per_turno(turno, mese)
            cell_progr = ws.cell(row=riga, column=giorni_mese + 4, value=progressivo_turno)

            # Aggiungi bordi alle colonne finali
            thick_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            cell_turno.border = thick_border
            cell_gg.border = thick_border
            cell_progr.border = thick_border
            cell_turno.alignment = Alignment(horizontal='center')
            cell_gg.alignment = Alignment(horizontal='center')
            cell_progr.alignment = Alignment(horizontal='center')

            riga += 1

            # Aggiungi 2 righe bianche DOPO il turno 46
            if turno == 46:
                riga += 2

        # Larghezza colonne
        ws.column_dimensions['A'].width = 8
        for col in range(2, giorni_mese + 2):
            ws.column_dimensions[get_column_letter(col)].width = 4
        ws.column_dimensions[get_column_letter(giorni_mese + 2)].width = 8
        ws.column_dimensions[get_column_letter(giorni_mese + 3)].width = 6
        ws.column_dimensions[get_column_letter(giorni_mese + 4)].width = 8

    def _calcola_progressivo_per_turno(self, turno: int, mese: int) -> int:
        """Calcola il progressivo dei giorni lavorativi per uno specifico turno fino al mese corrente (incluso)"""
        progressivo = 0

        for m in range(1, mese + 1):
            giorni_mese = calendar.monthrange(self.anno_target, m)[1]
            for giorno in range(1, giorni_mese + 1):
                data = date(self.anno_target, m, giorno)
                codice = self.calendario.get(data, {}).get(turno, '-')
                # Conta giorni lavorativi inclusi G e F
                if codice in ['A', 'B', 'C', 'G'] or codice.startswith('F'):
                    progressivo += 1

        return progressivo

    def _calcola_progressivo_precedente(self, mese: int) -> int:
        """Calcola il progressivo dei giorni lavorativi dei mesi precedenti (DEPRECATO - mantenuto per compatibilità)"""
        progressivo = 0

        for m in range(1, mese):
            giorni_mese = calendar.monthrange(self.anno_target, m)[1]
            for giorno in range(1, giorni_mese + 1):
                data = date(self.anno_target, m, giorno)
                # Conta per il primo turno (tutti hanno stesso numero di giorni lavorativi)
                codice = self.calendario.get(data, {}).get(41, '-')
                if codice in ['A', 'B', 'C', 'G'] or codice.startswith('F'):
                    progressivo += 1

        return progressivo

    def _formatta_cella(self, cell, codice: str, data: date, turno: int):
        """Applica formattazione alla cella"""

        # Colore di sfondo - priorità alle festività/weekend, poi ai codici
        if is_festivo(data) or (data.weekday() == 6):  # Domenica o festivo
            # Rosso per domeniche e festivi
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)  # Testo bianco
        elif data.weekday() == 5:  # Sabato
            # Azzurro scuro per sabati
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)  # Testo bianco
        elif codice == 'G' and turno != 46:
            # Verde chiaro per G, ECCETTO turno 46
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif codice.startswith('F'):
            # Giallo chiaro per ferie
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        # Bordi
        thin_border = Side(style='thin', color='000000')
        cell.border = Border(left=thin_border, right=thin_border,
                           top=thin_border, bottom=thin_border)

    def genera_report(self, verifica: Dict, output_path: str):
        """Genera un report testuale della verifica"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("=" * 70 + "\n")
            f.write(f"REPORT VERIFICA TURNI {self.anno_target}\n")
            f.write("=" * 70 + "\n\n")

            f.write(f"Giorni verificati OK: {verifica['statistiche']['ok']}\n")
            f.write(f"Giorni con anomalie: {verifica['statistiche']['anomalie']}\n\n")

            if verifica['anomalie']:
                f.write("ANOMALIE RILEVATE:\n")
                f.write("-" * 70 + "\n")
                for anomalia in verifica['anomalie']:
                    f.write(f"Data: {anomalia['data'].strftime('%d/%m/%Y')} - ")
                    f.write(f"A={anomalia['A']}, B={anomalia['B']}, C={anomalia['C']}, G={anomalia['G']}\n")
            else:
                f.write("Nessuna anomalia rilevata.\n")

            f.write("\n" + "=" * 70 + "\n")
            f.write("NOTA: Questo è un output da verificare manualmente.\n")
            f.write("=" * 70 + "\n")

# ============================================================================
# MAIN
# ============================================================================

def main():
    """Funzione principale"""

    print("=" * 70)
    print("GENERATORE TURNI RAFFINERIA")
    print("=" * 70)

    # Configurazione
    file_template = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\TURNO COMPLETO 2024.xls"
    anno_target = 2025

    output_dir = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\PROGETTI_PYTHON_OFFICE\SHIFT_MANAGERV1.0"
    output_excel = os.path.join(output_dir, f"TURNO_COMPLETO_{anno_target}.xlsx")
    output_report = os.path.join(output_dir, "report_verifica.txt")

    # Verifica esistenza file template
    if not os.path.exists(file_template):
        print(f"\nERRORE: File template non trovato: {file_template}")
        return

    # Creazione generatore
    generatore = GeneratoreTurni(file_template, anno_target)

    # Step 1: Lettura template
    if not generatore.leggi_template():
        print("\nERRORE: Impossibile leggere il template")
        return

    # Step 2: Generazione calendario
    generatore.genera_calendario()

    # Step 3: Applicazione ferie
    generatore.applica_ferie()

    # Step 4: Bilanciamento giorni lavorativi (target 231)
    generatore.bilancia_giorni_lavorativi(231)

    # Step 5: Verifica copertura
    verifica = generatore.verifica_copertura()

    # Step 5: Generazione output
    generatore.genera_excel(output_excel)
    generatore.genera_report(verifica, output_report)

    print("\n" + "=" * 70)
    print(f"File generato: {output_excel}")
    print(f"Report generato: {output_report}")
    print("=" * 70)
    print("\nRICHIESTA VERIFICA MANUALE")
    print("=" * 70)

if __name__ == "__main__":
    main()

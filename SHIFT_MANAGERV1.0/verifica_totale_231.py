"""
Verifica che tutti i turni abbiano esattamente 231 giorni lavorativi
"""
import openpyxl

file_generato = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\PROGETTI_PYTHON_OFFICE\SHIFT_MANAGERV1.0\TURNO_COMPLETO_2025.xlsx"

wb = openpyxl.load_workbook(file_generato, data_only=True)

print("="*80)
print("VERIFICA TOTALE 231 GIORNI LAVORATIVI")
print("="*80)

# Leggi progressivi da DICEMBRE
ws_dic = wb['DICEMBRE']

turni = [41, 42, 43, 44, 45, 47, 48, 49, 50, 51]

print("\nProgressivi finali (da colonna 'Progr.' in DICEMBRE):")
print("-" * 80)

problemi = []

for row in range(3, ws_dic.max_row + 1):
    turno = ws_dic.cell(row, 1).value
    if turno in turni:
        # Il progressivo è nell'ultima colonna
        progressivo = ws_dic.cell(row, ws_dic.max_column).value

        status = " OK" if progressivo == 231 else " ERRORE!"

        print(f"Turno {turno}: {progressivo} giorni{status}")

        if progressivo != 231:
            problemi.append((turno, progressivo))

print("\n" + "="*80)
if problemi:
    print(f"ATTENZIONE: {len(problemi)} turni NON hanno 231 giorni!")
    for turno, prog in problemi:
        diff = 231 - prog
        print(f"  Turno {turno}: {prog} giorni (mancano {diff})")
else:
    print("SUCCESSO: Tutti i turni hanno esattamente 231 giorni!")
print("="*80)

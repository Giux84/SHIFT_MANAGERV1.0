"""
Verifica le tre modifiche richieste:
1. G days dopo il giorno 14
2. Turno 46 bilanciato (no mesi estivi)
3. Ferie sempre gialle
"""
import openpyxl

file_generato = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\PROGETTI_PYTHON_OFFICE\SHIFT_MANAGERV1.0\TURNO_COMPLETO_2025.xlsx"

wb = openpyxl.load_workbook(file_generato)

print("="*80)
print("VERIFICA MODIFICHE")
print("="*80)

# VERIFICA 1: G days solo dopo il giorno 14
print("\n1. VERIFICA G DAYS DOPO GIORNO 14:")
print("-" * 80)

mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'OTTOBRE', 'NOVEMBRE']
turni = [41, 42, 43, 44, 45, 47, 48, 49, 50, 51]

g_prima_del_15 = []

for mese in mesi:
    if mese not in wb.sheetnames:
        continue

    ws = wb[mese]

    for row in range(3, ws.max_row + 1):
        turno = ws.cell(row, 1).value
        if turno not in turni:
            continue

        for giorno in range(1, 15):  # Giorni 1-14
            col = giorno + 1
            val = ws.cell(row, col).value
            if val == 'G':
                g_prima_del_15.append(f"{mese} - Turno {turno} - Giorno {giorno}")

if g_prima_del_15:
    print("  ERRORE: Trovati G prima del giorno 15:")
    for item in g_prima_del_15:
        print(f"    {item}")
else:
    print("  OK - Tutti i G sono dopo il giorno 14")

# VERIFICA 2: Turno 46 bilanciato (no G in giugno, luglio, agosto, settembre)
print("\n2. VERIFICA TURNO 46:")
print("-" * 80)

turno46_g = {}
for mese in ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
             'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']:
    if mese not in wb.sheetnames:
        continue

    ws = wb[mese]

    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 1).value == 46:
            g_count = 0
            for col in range(2, 33):
                if ws.cell(row, col).value == 'G':
                    g_count += 1
            turno46_g[mese] = g_count
            break

print("  Distribuzione G per Turno 46:")
total_g = 0
errori_46 = []
for mese, count in turno46_g.items():
    status = ""
    if count > 0:
        total_g += count
        # Verifica che non ci siano G nei mesi estivi
        if mese in ['GIUGNO', 'LUGLIO', 'AGOSTO', 'SETTEMBRE']:
            status = " ERRORE!"
            errori_46.append(f"{mese} ha {count} G (non dovrebbe averne)")
    print(f"    {mese}: {count} G{status}")

print(f"  Totale G per Turno 46: {total_g}")

if errori_46:
    print("  ERRORI rilevati:")
    for err in errori_46:
        print(f"    {err}")
else:
    print("  OK - Nessun G nei mesi estivi")

# VERIFICA 3: Controllo progressivo turno 46
ws_dic = wb['DICEMBRE']
for row in range(3, ws_dic.max_row + 1):
    if ws_dic.cell(row, 1).value == 46:
        prog = ws_dic.cell(row, ws_dic.max_column).value
        print(f"  Progressivo Dicembre Turno 46: {prog} giorni")
        if prog == 231:
            print("  OK - Turno 46 ha 231 giorni totali")
        else:
            print(f"  ATTENZIONE - Turno 46 ha {prog} giorni invece di 231")
        break

print("\n" + "="*80)
print("NOTA: La verifica del colore giallo delle ferie va fatta aprendo")
print("      il file Excel e controllando visivamente le celle con ferie")
print("      che cadono di sabato/domenica.")
print("="*80)

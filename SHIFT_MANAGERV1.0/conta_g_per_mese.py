import openpyxl
import sys

file_excel = sys.argv[1] if len(sys.argv) > 1 else "TURNO_COMPLETO_2026.xlsx"
wb = openpyxl.load_workbook(file_excel, data_only=True)

mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
       'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']

print("CONTEGGIO GIORNI G PER TURNO E MESE:")
print("="*70)

for mese in mesi:
    if mese not in wb.sheetnames:
        continue

    ws = wb[mese]
    print(f"\n{mese}:")

    for row in range(3, 16):
        turno = ws.cell(row, 1).value
        if not turno or not str(turno).isdigit():
            continue

        g_count = 0
        for col in range(2, 33):  # Max 31 giorni
            cell_value = ws.cell(row, col).value
            if cell_value == 'G':
                g_count += 1

        if g_count > 0:
            status = " MULTIPLI!" if g_count > 1 else ""
            print(f"  Turno {turno}: {g_count} G{status}")

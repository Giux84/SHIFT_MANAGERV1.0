"""
Script per analizzare le differenze tra file generato e file ufficiale
"""

import pandas as pd
import openpyxl

# File paths
file_ufficiale = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\TURNO COMPLETO 2025 con modifiche da CCNL.xls"
file_generato = r"D:\Users\gcaravel\OneDrive - ram.it\Desktop\TURNO_COMPLETO_2025.xlsx"

print("="*80)
print("ANALISI DIFFERENZE FILE TURNI 2025")
print("="*80)

# Leggi file ufficiale (.xls con xlrd)
print("\nLettura file UFFICIALE (.xls)...")
try:
    excel_uff = pd.ExcelFile(file_ufficiale, engine='xlrd')
    print(f"Fogli disponibili: {excel_uff.sheet_names}")
except Exception as e:
    print(f"Errore lettura file ufficiale: {e}")
    exit(1)

# Leggi file generato (.xlsx)
print("\nLettura file GENERATO (.xlsx)...")
wb_gen = openpyxl.load_workbook(file_generato, data_only=True)

# Analizza GENNAIO per tutti i turni
mese = "GENNAIO"
print(f"\n{'='*80}")
print(f"ANALISI MESE: {mese}")
print(f"{'='*80}")

df_uff = pd.read_excel(excel_uff, sheet_name=mese, header=1)
ws_gen = wb_gen[mese]

# Trova i turni
turni = [41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51]

for turno in turni:
    print(f"\n--- TURNO {turno} ---")

    # Trova riga nel file ufficiale
    riga_uff = df_uff[df_uff.iloc[:, 0] == turno]
    if riga_uff.empty:
        print("  Non trovato nel file ufficiale")
        continue

    # Estrai primi 10 giorni dal file ufficiale
    giorni_uff = []
    for col in range(1, 11):  # Colonne 1-10 (giorni 1-10)
        val = riga_uff.iloc[0, col]
        if pd.isna(val):
            giorni_uff.append('-')
        else:
            giorni_uff.append(str(val).strip())

    # Estrai primi 10 giorni dal file generato
    # Trova la riga del turno
    riga_gen = None
    for r in range(3, ws_gen.max_row + 1):
        if ws_gen.cell(r, 1).value == turno:
            riga_gen = r
            break

    if riga_gen:
        giorni_gen = []
        for col in range(2, 12):  # Colonne 2-11 (giorni 1-10)
            val = ws_gen.cell(riga_gen, col).value
            if val is None:
                giorni_gen.append('-')
            else:
                giorni_gen.append(str(val).strip())
    else:
        print("  Non trovato nel file generato")
        continue

    # Confronta
    print(f"  UFFICIALE: {' '.join(giorni_uff)}")
    print(f"  GENERATO:  {' '.join(giorni_gen)}")

    # Evidenzia differenze
    differenze = []
    for i, (u, g) in enumerate(zip(giorni_uff, giorni_gen), 1):
        if u != g:
            differenze.append(f"Giorno {i}: '{u}' vs '{g}'")

    if differenze:
        print(f"  DIFFERENZE: {', '.join(differenze)}")
    else:
        print("  OK - Identici")

# Conta i G per ogni turno in tutti i mesi
print(f"\n{'='*80}")
print("CONTEGGIO GIORNI G PER TURNO (tutti i mesi)")
print(f"{'='*80}")

mesi = ['GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
       'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE']

for turno in turni:
    if turno == 46:
        continue  # Skip turno 46 che ha pattern diverso

    g_uff = 0
    g_gen = 0

    for mese in mesi:
        # File ufficiale
        try:
            df = pd.read_excel(excel_uff, sheet_name=mese, header=1)
            riga = df[df.iloc[:, 0] == turno]
            if not riga.empty:
                for col in range(1, len(riga.columns)):
                    val = riga.iloc[0, col]
                    if val == 'G':
                        g_uff += 1
        except:
            pass

        # File generato
        if mese in wb_gen.sheetnames:
            ws = wb_gen[mese]
            for r in range(3, ws.max_row + 1):
                if ws.cell(r, 1).value == turno:
                    for col in range(2, 33):
                        if ws.cell(r, col).value == 'G':
                            g_gen += 1
                    break

    print(f"Turno {turno}: UFFICIALE={g_uff} G, GENERATO={g_gen} G")

# Progressivi finali
print(f"\n{'='*80}")
print("PROGRESSIVI DICEMBRE")
print(f"{'='*80}")

df_dic_uff = pd.read_excel(excel_uff, sheet_name='DICEMBRE', header=1)
ws_dic_gen = wb_gen['DICEMBRE']

for turno in turni:
    # Ufficiale
    riga_uff = df_dic_uff[df_dic_uff.iloc[:, 0] == turno]
    prog_uff = "N/A"
    if not riga_uff.empty:
        # Progressivo è nell'ultima colonna
        prog_uff = riga_uff.iloc[0, -1]

    # Generato
    prog_gen = "N/A"
    for r in range(3, ws_dic_gen.max_row + 1):
        if ws_dic_gen.cell(r, 1).value == turno:
            prog_gen = ws_dic_gen.cell(r, ws_dic_gen.max_column).value
            break

    print(f"Turno {turno}: UFFICIALE={prog_uff}, GENERATO={prog_gen}")

print(f"\n{'='*80}")
print("ANALISI COMPLETATA")
print(f"{'='*80}")
